"""
Excel parser service for Inventor BOM files.

Handles parsing of Autodesk Inventor BOM exports in Excel format.
"""
import base64
import io
import re
import logging
from collections import defaultdict
from zipfile import ZipFile
import xml.etree.ElementTree as ET

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _logger.warning("openpyxl not installed. Excel import will not work.")


# XML namespaces used in Excel XLSX files
XLSX_NS = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}


# Column mapping for Inventor BOM (0-indexed)
COL_ITEM = 0          # A: Item (skip)
COL_PART_NUMBER = 1   # B: Part Number (product name)
COL_THUMBNAIL = 2     # C: Thumbnail (image)
COL_BOM_STRUCTURE = 3 # D: BOM Structure (skip)
COL_UNIT_QTY = 4      # E: Unit QTY (skip)
COL_QTY = 5           # F: QTY (quantity)
COL_STOCK_NUMBER = 6  # G: Stock Number (skip)
COL_DESCRIPTION = 7   # H: Description
COL_REV = 8           # I: REV (skip)
COL_MASS = 9          # J: Mass (weight in kg)


def check_openpyxl():
    """Check if openpyxl is available."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "The 'openpyxl' library is required for Excel import. "
            "Please install it with: pip install openpyxl"
        )


def parse_mass_to_kg(mass_str):
    """
    Parse mass string to float kg value.
    
    Handles formats like:
    - "13,600 kg" (comma as decimal)
    - "13.600 kg" (dot as decimal)
    - "1 234,567 kg" (space as thousands separator)
    - "13600" (plain number)
    
    Returns weight in kg as float, or 0.0 if parsing fails.
    """
    if not mass_str:
        return 0.0
    
    if isinstance(mass_str, (int, float)):
        return float(mass_str)
    
    # Convert to string and clean
    mass_str = str(mass_str).strip().lower()
    
    # Remove 'kg' suffix
    mass_str = re.sub(r'\s*kg\s*$', '', mass_str)
    
    # Remove spaces (thousands separator)
    mass_str = mass_str.replace(' ', '')
    
    # Detect decimal separator
    # If both comma and dot exist, the last one is likely decimal
    has_comma = ',' in mass_str
    has_dot = '.' in mass_str
    
    if has_comma and has_dot:
        # Both present - assume last one is decimal
        if mass_str.rfind(',') > mass_str.rfind('.'):
            # Comma is decimal (European: 1.234,56)
            mass_str = mass_str.replace('.', '').replace(',', '.')
        else:
            # Dot is decimal (US: 1,234.56)
            mass_str = mass_str.replace(',', '')
    elif has_comma:
        # Only comma - treat as decimal
        mass_str = mass_str.replace(',', '.')
    # If only dot or neither, leave as is
    
    try:
        return float(mass_str)
    except ValueError:
        _logger.warning(f"Could not parse mass value: {mass_str}")
        return 0.0


def extract_images_by_row_from_zip(file_content, thumbnail_col=COL_THUMBNAIL):
    """
    Extract embedded images from XLSX file by parsing the ZIP structure directly.
    
    This method parses the drawing XML files inside the XLSX to properly map
    images to their anchor positions. This is more reliable than openpyxl's
    _images attribute which doesn't always load images correctly.
    
    Args:
        file_content: Binary content of the XLSX file
        thumbnail_col: 0-based column index for thumbnails (default: 2 = column C)
    
    Returns:
        dict: {excel_row_number (1-based): image_bytes}
    """
    images_by_row = {}
    
    try:
        with ZipFile(io.BytesIO(file_content), 'r') as zf:
            # Check if drawing files exist
            drawing_rels_path = 'xl/drawings/_rels/drawing1.xml.rels'
            drawing_path = 'xl/drawings/drawing1.xml'
            
            if drawing_rels_path not in zf.namelist():
                _logger.debug("No drawing relationships file found")
                return images_by_row
            
            if drawing_path not in zf.namelist():
                _logger.debug("No drawing XML file found")
                return images_by_row
            
            # Parse drawing relationships (maps rId to image file path)
            rels_xml = zf.read(drawing_rels_path)
            rels_root = ET.fromstring(rels_xml)
            
            rid_to_image_path = {}
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rid = rel.get('Id')
                target = rel.get('Target')
                if target and 'media/image' in target:
                    # Normalize path: ../media/image1.jpg -> xl/media/image1.jpg
                    normalized = target.replace('../', 'xl/').lstrip('/')
                    rid_to_image_path[rid] = normalized
            
            _logger.debug(f"Found {len(rid_to_image_path)} image relationships")
            
            # Parse drawing XML to get anchor positions
            drawing_xml = zf.read(drawing_path)
            drawing_root = ET.fromstring(drawing_xml)
            
            # Find all twoCellAnchor elements (image anchors)
            for anchor in drawing_root.findall('.//xdr:twoCellAnchor', XLSX_NS):
                from_elem = anchor.find('xdr:from', XLSX_NS)
                if from_elem is None:
                    continue
                
                row_elem = from_elem.find('xdr:row', XLSX_NS)
                col_elem = from_elem.find('xdr:col', XLSX_NS)
                
                if row_elem is None or col_elem is None:
                    continue
                
                row_0based = int(row_elem.text)
                col_0based = int(col_elem.text)
                
                # Filter: only images in the Thumbnail column
                if col_0based != thumbnail_col:
                    continue
                
                excel_row = row_0based + 1  # Convert to 1-based
                
                # Get the image reference (blip embed)
                blip = anchor.find('.//a:blip', XLSX_NS)
                if blip is None:
                    continue
                
                embed_rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                if not embed_rid or embed_rid not in rid_to_image_path:
                    continue
                
                image_path = rid_to_image_path[embed_rid]
                
                # Read image data from ZIP
                try:
                    image_data = zf.read(image_path)
                    images_by_row[excel_row] = image_data
                    _logger.debug(f"Extracted image for row {excel_row}: {len(image_data)} bytes")
                except KeyError:
                    _logger.warning(f"Image file not found in archive: {image_path}")
            
            _logger.info(f"Extracted {len(images_by_row)} images mapped to rows")
            
    except Exception as e:
        _logger.warning(f"Could not extract images from ZIP: {e}")
    
    return images_by_row


def extract_images_by_row(worksheet, thumbnail_col=COL_THUMBNAIL):
    """
    Extract embedded images from worksheet using openpyxl's _images.
    
    Note: This method may not work for all XLSX files. Consider using
    extract_images_by_row_from_zip() for more reliable extraction.
    
    Args:
        worksheet: openpyxl worksheet object
        thumbnail_col: 0-based column index for thumbnails (default: 2 = column C)
    
    Returns:
        dict: {excel_row_number (1-based): image_bytes}
    """
    images_by_row = {}
    
    if not hasattr(worksheet, '_images') or not worksheet._images:
        _logger.debug("No embedded images found via openpyxl _images")
        return images_by_row
    
    _logger.info(f"Found {len(worksheet._images)} embedded images in worksheet")
    
    for image in worksheet._images:
        try:
            anchor = image.anchor
            
            if hasattr(anchor, '_from'):
                row_0based = anchor._from.row
                col_0based = anchor._from.col
            else:
                continue
            
            if col_0based != thumbnail_col:
                continue
            
            excel_row = row_0based + 1
            
            if hasattr(image, '_data'):
                image_data = image._data()
            elif hasattr(image, 'ref'):
                if hasattr(image.ref, 'getvalue'):
                    image_data = image.ref.getvalue()
                elif hasattr(image.ref, 'read'):
                    image_data = image.ref.read()
                else:
                    continue
            else:
                continue
            
            if image_data:
                images_by_row[excel_row] = image_data
            
        except Exception as e:
            _logger.warning(f"Could not extract image: {e}")
    
    return images_by_row


def image_to_base64(image_bytes):
    """Convert image bytes to base64 string for Odoo."""
    if not image_bytes:
        return False
    return base64.b64encode(image_bytes).decode('utf-8')


def iter_bom_rows(file_content, filename=None):
    """
    Iterate over BOM rows from Excel file.
    
    Args:
        file_content: Binary content of the Excel file
        filename: Optional filename for logging
    
    Yields:
        dict with keys: part_number, quantity, description, weight_kg, image_base64, row_number
    """
    check_openpyxl()
    
    # Load workbook from bytes
    file_stream = io.BytesIO(file_content)
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active
    
    # Extract images using ZIP method (more reliable than openpyxl _images)
    images_by_row = extract_images_by_row_from_zip(file_content)
    
    # Fallback to openpyxl method if ZIP extraction found nothing
    if not images_by_row:
        _logger.debug("ZIP extraction found no images, trying openpyxl method")
        images_by_row = extract_images_by_row(ws)
    
    # Skip header row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        # Get cell values
        cells = list(row)
        if len(cells) < 10:
            continue
        
        part_number = cells[COL_PART_NUMBER].value
        if not part_number:
            continue
        
        qty_value = cells[COL_QTY].value
        try:
            quantity = float(qty_value) if qty_value else 1.0
        except (ValueError, TypeError):
            quantity = 1.0
        
        description = cells[COL_DESCRIPTION].value or ''
        mass_str = cells[COL_MASS].value
        weight_kg = parse_mass_to_kg(mass_str)
        
        # Get image for this row if available
        image_data = images_by_row.get(row_idx)
        image_base64 = image_to_base64(image_data) if image_data else False
        
        yield {
            'part_number': str(part_number).strip(),
            'quantity': quantity,
            'description': str(description).strip() if description else '',
            'weight_kg': weight_kg,
            'image_base64': image_base64,
            'row_number': row_idx,
        }


def aggregate_bom_rows(file_content, filename=None):
    """
    Parse BOM and aggregate duplicate part numbers.
    
    Returns:
        tuple: (aggregated_rows, duplicates_info)
        - aggregated_rows: list of dicts with summed quantities
        - duplicates_info: list of part numbers that had duplicates
    """
    rows_by_part = defaultdict(lambda: {
        'quantity': 0.0,
        'description': '',
        'weight_kg': 0.0,
        'image_base64': False,
        'occurrences': 0,
    })
    
    for row in iter_bom_rows(file_content, filename):
        part = row['part_number']
        data = rows_by_part[part]
        data['quantity'] += row['quantity']
        data['occurrences'] += 1
        # Keep first non-empty values
        if not data['description'] and row['description']:
            data['description'] = row['description']
        if not data['weight_kg'] and row['weight_kg']:
            data['weight_kg'] = row['weight_kg']
        if not data['image_base64'] and row['image_base64']:
            data['image_base64'] = row['image_base64']
    
    aggregated = []
    duplicates = []
    
    for part_number, data in rows_by_part.items():
        aggregated.append({
            'part_number': part_number,
            'quantity': data['quantity'],
            'description': data['description'],
            'weight_kg': data['weight_kg'],
            'image_base64': data['image_base64'],
        })
        if data['occurrences'] > 1:
            duplicates.append({
                'part_number': part_number,
                'occurrences': data['occurrences'],
                'total_qty': data['quantity'],
            })
    
    return aggregated, duplicates
